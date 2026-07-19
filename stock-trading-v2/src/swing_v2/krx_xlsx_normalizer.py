"""Conservative, local-only normalization of KRX Data Marketplace XLSX exports.

The source date is supplied from the filename/user assertion, not inferred from
workbook contents.  Outputs are therefore forward-only from ``as_of``.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
import hashlib
import json
from pathlib import Path
import re
from typing import Iterable
import warnings

from openpyxl import load_workbook


MAPPING_VERSION = "krx-xlsx-conservative-v1"
SOURCE = "KRX Data Marketplace XLSX export (filename-declared as_of; user supplied)"


@dataclass(frozen=True)
class NormalizationResult:
    record_count: int
    content_hash: str
    output_path: Path
    manifest_path: Path


def normalize_krx_xlsx(
    stocks_path: str | Path,
    etf_detail_path: str | Path,
    etf_basic_path: str | Path,
    output_path: str | Path,
    manifest_path: str | Path,
    *,
    as_of: date,
) -> NormalizationResult:
    """Create loader-compatible, immutable JSON and a source-hash manifest.

    All source paths are read-only.  ``as_of`` is explicitly caller supplied and
    each output record starts on it, preventing historical backfill.
    """
    if type(as_of) is not date:
        raise ValueError("as_of must be a plain date")
    sources = {
        "stocks": Path(stocks_path),
        "etf_detail": Path(etf_detail_path),
        "etf_basic": Path(etf_basic_path),
    }
    for path in sources.values():
        if not path.is_file() or path.suffix.lower() != ".xlsx":
            raise ValueError(f"source must be an existing XLSX file: {path}")

    source_hashes = {role: _file_hash(path) for role, path in sources.items()}
    provenance_hash = "sha256:" + hashlib.sha256(
        json.dumps(source_hashes, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    stock_records = _stock_records(_read_rows(sources["stocks"]), as_of, provenance_hash)
    etf_records = _etf_records(
        _read_rows(sources["etf_detail"]), _read_rows(sources["etf_basic"]), as_of, provenance_hash,
    )
    # The all-securities export may include ETF codes.  The joined ETF sources
    # are more specific, so they replace only the stock export's UNKNOWN row.
    by_symbol = {str(record["symbol"]): record for record in stock_records}
    for record in etf_records:
        symbol = str(record["symbol"])
        existing = by_symbol.get(symbol)
        if existing is not None and existing["asset_type"] != "UNKNOWN":
            raise ValueError(f"stock/ETF code collision is not conservatively resolvable: {symbol}")
        by_symbol[symbol] = record
    records = list(by_symbol.values())
    records.sort(key=lambda row: str(row["symbol"]))
    canonical = json.dumps(records, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    content_hash = "sha256:" + hashlib.sha256(canonical).hexdigest()
    normalized = {"format_version": 1, "records": records}
    output = Path(output_path)
    manifest = Path(manifest_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    manifest.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(normalized, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    manifest.write_text(json.dumps({
        "format_version": 1,
        "mapping_version": MAPPING_VERSION,
        "as_of": as_of.isoformat(),
        "as_of_basis": "filename/user supplied; not intrinsic workbook proof",
        "caveat": "Eligible only for signal dates on or after as_of; never use for historical membership.",
        "normalized_content_hash": content_hash,
        "normalized_content_hash_definition": "sha256 of canonical UTF-8 JSON records (sorted keys, compact separators)",
        "record_provenance_content_hash": provenance_hash,
        "record_provenance_content_hash_definition": "sha256 of canonical JSON mapping source role to source-file sha256",
        "sources": [{"role": role, "filename": path.name, "sha256": source_hashes[role]} for role, path in sources.items()],
    }, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return NormalizationResult(len(records), content_hash, output, manifest)


def _read_rows(path: Path) -> list[dict[str, str]]:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style", category=UserWarning)
        workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        # KRX exports carry an incorrect ``A1`` worksheet dimension despite
        # having rows; read-only openpyxl otherwise silently returns header-only.
        sheet.reset_dimensions()
        values = sheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            raise ValueError(f"XLSX has no header row: {path.name}")
        columns = [str(value).strip() if value is not None else "" for value in headers]
        return [
            {column: _text(value) for column, value in zip(columns, row)}
            for row in values if any(value is not None and str(value).strip() for value in row)
        ]
    finally:
        workbook.close()


def _stock_records(rows: Iterable[dict[str, str]], as_of: date, provenance_hash: str) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for row in rows:
        symbol = _krx_short_code(row.get("단축코드", ""))
        if symbol is None:
            raise ValueError("stock XLSX has missing or invalid 단축코드")
        category = row.get("소속부", "")
        foreign_red_flag = "외국" in " ".join(row.values())
        if "SPAC" in category.upper():
            asset_type, flags = "SPAC", []
        elif row.get("증권구분") == "주권" and row.get("주식종류") == "보통주" and not foreign_red_flag:
            asset_type = "STOCK"
            flags = ["MANAGEMENT_ISSUE"] if "관리종목" in category else []
        else:
            asset_type, flags = "UNKNOWN", []
        records.append(_record(symbol, asset_type, flags, None, as_of, provenance_hash))
    return records


def _etf_records(
    detail_rows: Iterable[dict[str, str]], basic_rows: Iterable[dict[str, str]], as_of: date, provenance_hash: str,
) -> list[dict[str, object]]:
    detail_by_code = _unique_rows_by_code(detail_rows, "단축코드", "ETF detail")
    basic_by_code = _unique_rows_by_code(basic_rows, "종목코드", "ETF basic")
    if set(detail_by_code) != set(basic_by_code):
        raise ValueError("ETF detail/basic code mismatch")
    records: list[dict[str, object]] = []
    for symbol in sorted(detail_by_code):
        detail, basic = detail_by_code[symbol], basic_by_code[symbol]
        taxonomy = basic.get("분류체계", "")
        taxonomy_allowed = taxonomy == "주식-시장대표" or taxonomy.startswith("주식-업종섹터")
        market, asset, tracking = (detail.get(field, "") for field in ("기초시장분류", "기초자산분류", "추적배수"))
        flags: list[str] = []
        tracking_folded = tracking.casefold()
        if "레버리지" in tracking or "leverage" in tracking_folded:
            flags.append("ETF_LEVERAGED")
        if "인버스" in tracking or "inverse" in tracking_folded:
            flags.append("ETF_INVERSE")
        if market != "국내":
            flags.append("ETF_FOREIGN_INDEX")
        eligible_classification = market == "국내" and asset == "주식" and tracking in {"일반", "normal", "NORMAL"} and taxonomy_allowed
        # Retain an ETF asset type whenever the classification taxonomy is clear;
        # the flags make non-eligible products auditably excluded.
        asset_type = "ETF" if taxonomy_allowed else "UNKNOWN"
        exposure = "DOMESTIC_INDEX_OR_SECTOR" if taxonomy_allowed else None
        if eligible_classification:
            asset_type, exposure = "ETF", "DOMESTIC_INDEX_OR_SECTOR"
        records.append(_record(symbol, asset_type, flags, exposure, as_of, provenance_hash))
    return records


def _unique_rows_by_code(rows: Iterable[dict[str, str]], code_field: str, label: str) -> dict[str, dict[str, str]]:
    by_code: dict[str, dict[str, str]] = {}
    for row in rows:
        code = _krx_short_code(row.get(code_field, ""))
        if code is None:
            raise ValueError(f"{label} XLSX has missing or invalid {code_field}")
        if code in by_code:
            raise ValueError(f"{label} XLSX has duplicate code: {code}")
        by_code[code] = row
    return by_code


def _krx_short_code(value: str) -> str | None:
    code = value.strip().upper()
    if code.isdigit():
        code = code.zfill(6)
    return code if re.fullmatch(r"[0-9A-Z]{6}", code) else None


def _record(symbol: str, asset_type: str, flags: list[str], etf_exposure: str | None, as_of: date, provenance_hash: str) -> dict[str, object]:
    return {"symbol": symbol, "asset_type": asset_type, "effective_from": as_of.isoformat(), "effective_to": None,
            "flags": flags, "etf_exposure": etf_exposure, "source": SOURCE, "version": MAPPING_VERSION,
            "content_hash": provenance_hash, "as_of": as_of.isoformat()}


def _file_hash(path: Path) -> str:
    return "sha256:" + hashlib.sha256(path.read_bytes()).hexdigest()


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
